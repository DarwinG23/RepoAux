import openpyxl
from tabulate import tabulate
import ast

class Read:
    def __init__(self, file):
        self.__file = file
        self.__info = []

    @property
    def _info(self):
        return self.__info

    @_info.setter
    def _info(self, value):
        self.__info = value


    @property
    def _ruta(self):
        return self.__ruta

    @_ruta.setter
    def _ruta(self, value):
        self.__ruta = value

    
    def leer_archivo(self):
        exel_dataframe =  openpyxl.load_workbook(self.__file)
        
        dataframe =  exel_dataframe.active
        
        data = []
        
        for row in range(2,dataframe.max_row+1):
            _row = [row,]
            for col in range(1, dataframe.max_column+1):
                cell_value = dataframe.cell(row=row, column=col).value
                if isinstance(cell_value, (float)) and col != 4:
                    cell_value =   f"{cell_value:.0f}"
                if col == 5:
                    cell_value =   f"{cell_value:.2f}"
               
                _row.append(cell_value)
                    
            data.append(_row)
        
        self.__info = data 
        
    def info_to_dict(self):
        data = []
        for row in self.__info:
            _row = {
                "cedula": row[1],
                "nombres": row[2],
                "apellidos": row[3],
                "notas": row[4]
            }
            data.append(_row)
        return data
    

    def imprimir(self):
        if len(self.__info) == 0:
            print("Hubo un error al leer el archivo")
            return 
        headers = ["Nro","cedula","nombres", "apellidos", "notas"]
        print(tabulate(self.__info, headers=headers, tablefmt="fancy_grid"))
        
    
    def str_to_dict(self, data):
       data = ast.literal_eval(data)
       return data
